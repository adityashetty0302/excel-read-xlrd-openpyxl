import os
from openpyxl import load_workbook

if __name__ == '__main__':

    # cwd=os.getcwd()
    # print(cwd)

    os.chdir("C:\\Users\\Aditya.Shetty\\Desktop")
    # print(os.listdir('.'))

    wb = load_workbook('t1.xlsx')
    # print(wb.get_sheet_names())

    sheet = wb.get_sheet_by_name('Sheet1')
    # print(sheet.title)

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(i, j).value, "\t")
        print("\n")

    for cellObj in sheet['A1':'D5']:
        with open("Output.txt", "a") as text_file:
            for cell in cellObj:
                text_file.write(str(cell.value))
                text_file.write("\t")
            text_file.write("\n")
